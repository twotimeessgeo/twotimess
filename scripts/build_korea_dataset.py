from __future__ import annotations

import csv
import io
import json
import re
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_JSON = ROOT / "data" / "korea-climate-data.json"
OUTPUT_JS = ROOT / "data" / "korea-climate-data.js"
EXISTING_DATASET_JSON = ROOT / "data" / "korea-climate-data.json"

MONTHS = [f"{index}월" for index in range(1, 13)]
KMA_DOWNLOAD_URL = "https://data.kma.go.kr/download/downloadNormDataFile.do"
KMA_STATION_DOWNLOAD_URL = "https://data.kma.go.kr/tmeta/stn/selectStnListDownload.do"
KMA_STATION_DETAIL_URL = "https://data.kma.go.kr/tmeta/stn/selectStnDetail.do?pgmNo=82&isSelectStn=Y&stdStnNo={station_id}"

SOUTH_NORMALS_FILE = "average30yearsKorea_1991_month.xlsx"
SOUTH_NORMALS_REAL_NAME = "우리나라기후평년(월별)_1991.xlsx"
NORTH_NORMALS_FILE = "average30yearsNKorea_1991_month.xlsx"
NORTH_NORMALS_REAL_NAME = "북한기후평년(월별)_1991.xlsx"

ZONE_ORDER = [
    "전체",
    "수도권·서해",
    "강원 영서",
    "강원 영동",
    "충청",
    "호남",
    "영남",
    "제주",
    "북서부",
    "북동부",
    "북중부",
    "북남부",
]
NATION_ORDER = ["전체", "남한", "북한"]
DEFAULT_SAMPLE_NAMES = ["서울", "강릉", "평양", "제주"]


@dataclass(frozen=True)
class StationConfig:
    station_id: int
    display_name: str
    official_name: str
    nation: str
    zone: str
    aliases: tuple[str, ...] = ()


SOUTH_STATIONS = [
    StationConfig(102, "백령도", "백령도", "남한", "수도권·서해"),
    StationConfig(108, "서울", "서울", "남한", "수도권·서해"),
    StationConfig(98, "동두천", "동두천", "남한", "수도권·서해"),
    StationConfig(112, "인천", "인천", "남한", "수도권·서해"),
    StationConfig(99, "파주", "파주", "남한", "수도권·서해"),
    StationConfig(119, "수원", "수원", "남한", "수도권·서해"),
    StationConfig(101, "춘천", "춘천", "남한", "강원 영서"),
    StationConfig(95, "철원", "철원", "남한", "강원 영서"),
    StationConfig(114, "원주", "원주", "남한", "강원 영서"),
    StationConfig(121, "영월", "영월", "남한", "강원 영서"),
    StationConfig(90, "속초", "속초", "남한", "강원 영동"),
    StationConfig(100, "대관령", "대관령", "남한", "강원 영동"),
    StationConfig(106, "동해", "동해", "남한", "강원 영동"),
    StationConfig(105, "강릉", "강릉", "남한", "강원 영동"),
    StationConfig(129, "서산", "서산", "남한", "충청"),
    StationConfig(131, "청주", "청주", "남한", "충청"),
    StationConfig(133, "대전", "대전", "남한", "충청"),
    StationConfig(127, "충주", "충주", "남한", "충청"),
    StationConfig(135, "추풍령", "추풍령", "남한", "충청"),
    StationConfig(140, "군산", "군산", "남한", "호남"),
    StationConfig(864, "전주", "전주(완산)", "남한", "호남", ("전주(완산)",)),
    StationConfig(172, "고창", "고창", "남한", "호남"),
    StationConfig(156, "광주", "광주", "남한", "호남"),
    StationConfig(165, "목포", "목포", "남한", "호남"),
    StationConfig(169, "흑산도", "흑산도", "남한", "호남"),
    StationConfig(170, "완도", "완도", "남한", "호남"),
    StationConfig(168, "여수", "여수", "남한", "호남"),
    StationConfig(136, "안동", "안동", "남한", "영남"),
    StationConfig(137, "상주", "상주", "남한", "영남"),
    StationConfig(138, "포항", "포항", "남한", "영남"),
    StationConfig(130, "울진", "울진", "남한", "영남"),
    StationConfig(152, "울산", "울산", "남한", "영남"),
    StationConfig(155, "창원", "창원", "남한", "영남"),
    StationConfig(192, "진주", "진주", "남한", "영남"),
    StationConfig(159, "부산", "부산", "남한", "영남"),
    StationConfig(860, "대구", "대구(신암)", "남한", "영남", ("대구(신암)",)),
    StationConfig(162, "통영", "통영", "남한", "영남"),
    StationConfig(115, "울릉도", "울릉도", "남한", "영남"),
    StationConfig(185, "고산", "고산", "남한", "제주"),
    StationConfig(188, "성산", "성산", "남한", "제주"),
    StationConfig(184, "제주", "제주", "남한", "제주"),
    StationConfig(189, "서귀포", "서귀포", "남한", "제주"),
]

NORTH_STATIONS = [
    StationConfig(3, "선봉", "선봉", "북한", "북동부"),
    StationConfig(5, "삼지연", "삼지연", "북한", "북중부"),
    StationConfig(8, "청진", "청진", "북한", "북동부"),
    StationConfig(14, "중강", "중강", "북한", "북중부"),
    StationConfig(16, "혜산", "혜산", "북한", "북중부"),
    StationConfig(20, "강계", "강계", "북한", "북중부"),
    StationConfig(22, "풍산", "풍산", "북한", "북중부"),
    StationConfig(25, "김책", "김책", "북한", "북동부"),
    StationConfig(28, "수풍", "수풍", "북한", "북서부"),
    StationConfig(31, "장진", "장진", "북한", "북중부"),
    StationConfig(35, "신의주", "신의주", "북한", "북서부"),
    StationConfig(37, "구성", "구성", "북한", "북서부"),
    StationConfig(39, "희천", "희천", "북한", "북중부"),
    StationConfig(41, "함흥", "함흥", "북한", "북동부"),
    StationConfig(46, "신포", "신포", "북한", "북동부"),
    StationConfig(50, "안주", "안주", "북한", "북서부"),
    StationConfig(52, "양덕", "양덕", "북한", "북중부"),
    StationConfig(55, "원산", "원산", "북한", "북동부"),
    StationConfig(58, "평양", "평양", "북한", "북서부"),
    StationConfig(60, "남포", "남포", "북한", "북서부"),
    StationConfig(61, "장전", "장전", "북한", "북동부"),
    StationConfig(65, "사리원", "사리원", "북한", "북남부"),
    StationConfig(67, "신계", "신계", "북한", "북남부"),
    StationConfig(68, "용연", "용연", "북한", "북남부"),
    StationConfig(69, "해주", "해주", "북한", "북남부"),
    StationConfig(70, "개성", "개성", "북한", "북남부"),
    StationConfig(75, "평강", "평강", "북한", "북남부"),
]


def request_bytes(url: str, data: dict[str, str] | None = None) -> bytes:
    encoded = None
    if data is not None:
        encoded = urllib.parse.urlencode(data).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=encoded,
        headers={
            "User-Agent": "TWOTIMESS Climate Machine Dataset Builder/2.0",
            "Accept": "*/*",
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def download_normals_workbook(dist_file_name: str, real_file_name: str, target_path: Path) -> None:
    payload = request_bytes(
        KMA_DOWNLOAD_URL,
        {
            "distFileName": dist_file_name,
            "realFileName": real_file_name,
        },
    )
    target_path.write_bytes(payload)


def load_workbook(path: Path):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def parse_south_monthly_sheet(workbook, sheet_name: str) -> dict[int, list[float]]:
    rows: dict[int, list[float]] = {}
    worksheet = workbook[sheet_name]
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        station_id = row[0]
        if not station_id:
            continue
        values = [round_number(value) for value in row[2:14]]
        rows[int(station_id)] = values
    return rows


def parse_south_class_days(workbook) -> tuple[dict[int, list[float]], dict[int, list[float]]]:
    below_zero: dict[int, list[float]] = {}
    above_twenty_five: dict[int, list[float]] = {}
    worksheet = workbook["33.계급별일수"]
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        station_id = row[0]
        if not station_id:
            continue
        element = str(row[2] or "").strip()
        grade = str(row[3] or "").strip()
        if element != "일최저기온":
            continue
        values = [round_number(value) for value in row[4:16]]
        if grade == "<0":
            below_zero[int(station_id)] = values
        elif grade == ">=25":
            above_twenty_five[int(station_id)] = values
    return below_zero, above_twenty_five


def parse_north_monthly_normals(workbook) -> tuple[dict[int, list[float]], dict[int, list[float]]]:
    temperatures: dict[int, list[float]] = {}
    precipitation: dict[int, list[float]] = {}
    worksheet = workbook["월평년값"]
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        station_id = row[0]
        month = row[2]
        if not station_id or not month:
            continue
        month_index = int(month) - 1
        station_key = int(station_id)
        temperatures.setdefault(station_key, [0.0] * 12)[month_index] = round_number(row[5])
        precipitation.setdefault(station_key, [0.0] * 12)[month_index] = round_number(row[12])
    return temperatures, precipitation


def parse_north_class_days(workbook) -> tuple[dict[int, list[float]], dict[int, list[float]]]:
    below_zero: dict[int, list[float]] = {}
    above_twenty_five: dict[int, list[float]] = {}
    worksheet = workbook["월별_계급별일수"]
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        station_id = row[0]
        month = row[2]
        if not station_id or not month:
            continue
        month_index = int(month) - 1
        station_key = int(station_id)
        below_zero.setdefault(station_key, [0.0] * 12)[month_index] = round_number(row[17])
        above_twenty_five.setdefault(station_key, [0.0] * 12)[month_index] = round_number(row[19])
    return below_zero, above_twenty_five


def download_station_metadata(station_ids: list[int]) -> dict[int, dict[str, float]]:
    records: dict[int, dict[str, float]] = {}
    existing_records = load_existing_station_metadata()

    try:
        payload = request_bytes(
            KMA_STATION_DOWNLOAD_URL,
            {
                "fileType": "csv",
                "pageIndex": "1",
                "schListCnt": str(len(station_ids)),
                "mddlClssCd": "",
                "stnIds": ",".join(str(station_id) for station_id in station_ids),
                "serviceSe": "F00101",
                "txtStnNm": "",
                "txtElementNm": "",
                "dTreeId": "",
                "gTreeId": "",
                "mddlClssCdDiff": "",
                "pgmNo": "82",
            },
        )
        if not is_html_document(payload):
            text = decode_kma_csv(payload)
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if not row or not row[0].strip().isdigit():
                    continue
                station_id = int(row[0].strip())
                end_date = row[2].strip()
                start_date = row[1].strip()
                record = {
                    "latitude": float(row[6]),
                    "longitude": float(row[7]),
                    "elevationM": float(row[8]) if row[8].strip() else 0.0,
                    "_endDate": end_date,
                    "_startDate": start_date,
                }
                current = records.get(station_id)
                if current is None:
                    records[station_id] = record
                    continue
                current_is_open = current["_endDate"] == ""
                record_is_open = record["_endDate"] == ""
                if record_is_open and not current_is_open:
                    records[station_id] = record
                    continue
                if record_is_open == current_is_open and record["_startDate"] > current["_startDate"]:
                    records[station_id] = record
    except Exception:
        pass

    for record in records.values():
        record.pop("_endDate", None)
        record.pop("_startDate", None)

    for station_id in station_ids:
        if station_id not in records:
            try:
                records[station_id] = fetch_station_detail_metadata(station_id)
            except ValueError:
                if station_id not in existing_records:
                    raise
                records[station_id] = existing_records[station_id]

    return records


def fetch_station_detail_metadata(station_id: int) -> dict[str, float]:
    payload = request_bytes(KMA_STATION_DETAIL_URL.format(station_id=station_id))
    text = payload.decode("utf-8", errors="replace")

    coordinate_match = re.search(r"위도\s*:\s*([-0-9.]+).*?경도\s*:\s*([-0-9.]+)", text, re.S)
    elevation_match = re.search(r"해발고도\(m\)</th>\s*<td>\s*([-0-9.]+)\s*</td>", text, re.S)

    if coordinate_match is None or elevation_match is None:
        raise ValueError(f"Failed to parse KMA station detail page for station {station_id}")

    return {
        "latitude": float(coordinate_match.group(1)),
        "longitude": float(coordinate_match.group(2)),
        "elevationM": float(elevation_match.group(1)),
    }


def is_html_document(payload: bytes) -> bool:
    head = payload[:256].lstrip().lower()
    return head.startswith(b"<!doctype html") or head.startswith(b"<html")


def load_existing_station_metadata() -> dict[int, dict[str, float]]:
    if not EXISTING_DATASET_JSON.exists():
        return {}

    dataset = json.loads(EXISTING_DATASET_JSON.read_text(encoding="utf-8"))
    records: dict[int, dict[str, float]] = {}
    for region in dataset.get("regions", []):
        station_id = region.get("stationId")
        coordinates = region.get("coordinates") or {}
        if station_id is None:
            continue
        records[int(station_id)] = {
            "latitude": float(coordinates["latitude"]),
            "longitude": float(coordinates["longitude"]),
            "elevationM": float(region.get("elevationM", 0.0)),
        }
    return records


def decode_kma_csv(payload: bytes) -> str:
    for encoding in ("cp949", "euc-kr", "utf-8-sig", "utf-8"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("cp949", errors="replace")


def build_region(
    config: StationConfig,
    coordinates: dict[str, float],
    temperatures: dict[int, list[float]],
    precipitation: dict[int, list[float]],
    below_zero_days: dict[int, list[float]],
    above_twenty_five_days: dict[int, list[float]],
) -> dict:
    station_id = config.station_id
    monthly_temperature = temperatures[station_id]
    monthly_precipitation = precipitation[station_id]
    monthly_below_zero_days = below_zero_days.get(station_id, [0.0] * 12)
    monthly_above_twenty_five_days = above_twenty_five_days.get(station_id, [0.0] * 12)

    return {
        "id": f"kma-{station_id}",
        "stationId": station_id,
        "name": config.display_name,
        "officialName": config.official_name,
        "aliases": [config.display_name, config.official_name, *config.aliases],
        "nation": config.nation,
        "zone": config.zone,
        "months": MONTHS,
        "coordinates": {
            "latitude": round_coordinate(coordinates["latitude"]),
            "longitude": round_coordinate(coordinates["longitude"]),
        },
        "elevationM": round_number(coordinates["elevationM"]),
        "monthlyTemperatureC": monthly_temperature,
        "monthlyPrecipitationMm": monthly_precipitation,
        "monthlyColdDaysBelowZero": monthly_below_zero_days,
        "monthlyHotDaysAboveTwentyFiveMin": monthly_above_twenty_five_days,
        "annualMeanTemperatureC": average(monthly_temperature),
        "annualPrecipitationMm": round_number(sum(monthly_precipitation)),
        "annualColdDaysBelowZero": round_number(sum(monthly_below_zero_days)),
        "annualHotDaysAboveTwentyFiveMin": round_number(sum(monthly_above_twenty_five_days)),
        "source": {
            "type": "kma",
            "label": "기상청 기후평년값 / 지점정보",
            "period": "1991-2020",
            "note": "한국지리 버전은 기상청 1991-2020 월평년값과 기상청 지점정보를 기준으로 구성했습니다. 남북한 통일 비교를 위해 보조 지표는 일최저기온 계급별 일수를 사용합니다.",
        },
    }


def build_dataset() -> dict:
    download_normals_workbook(SOUTH_NORMALS_FILE, SOUTH_NORMALS_REAL_NAME, ROOT / "data" / SOUTH_NORMALS_FILE)
    download_normals_workbook(NORTH_NORMALS_FILE, NORTH_NORMALS_REAL_NAME, ROOT / "data" / NORTH_NORMALS_FILE)

    south_workbook = load_workbook(ROOT / "data" / SOUTH_NORMALS_FILE)
    north_workbook = load_workbook(ROOT / "data" / NORTH_NORMALS_FILE)

    south_temperatures = parse_south_monthly_sheet(south_workbook, "3.평균기온")
    south_precipitation = parse_south_monthly_sheet(south_workbook, "12.강수량")
    south_below_zero_days, south_above_twenty_five_days = parse_south_class_days(south_workbook)

    north_temperatures, north_precipitation = parse_north_monthly_normals(north_workbook)
    north_below_zero_days, north_above_twenty_five_days = parse_north_class_days(north_workbook)

    selected_station_ids = [station.station_id for station in SOUTH_STATIONS + NORTH_STATIONS]
    station_metadata = download_station_metadata(selected_station_ids)

    regions: list[dict] = []

    for config in SOUTH_STATIONS:
        regions.append(
            build_region(
                config,
                station_metadata[config.station_id],
                south_temperatures,
                south_precipitation,
                south_below_zero_days,
                south_above_twenty_five_days,
            )
        )

    for config in NORTH_STATIONS:
        regions.append(
            build_region(
                config,
                station_metadata[config.station_id],
                north_temperatures,
                north_precipitation,
                north_below_zero_days,
                north_above_twenty_five_days,
            )
        )

    regions.sort(key=lambda region: (ZONE_ORDER.index(region["zone"]), region["name"]))

    dataset = {
        "title": "한국 기후 통계",
        "subtitle": "기상청 1991-2020 평년값 기준",
        "months": MONTHS,
        "defaultSampleNames": DEFAULT_SAMPLE_NAMES,
        "nationOrder": NATION_ORDER,
        "zoneOrder": ZONE_ORDER,
        "comparisonPeriods": [
            {"id": "jan", "label": "1월", "monthIndexes": [0]},
            {"id": "aug", "label": "8월", "monthIndexes": [7]},
            {"id": "winter", "label": "겨울(12~2월)", "monthIndexes": [11, 0, 1]},
            {"id": "summer", "label": "여름(6~8월)", "monthIndexes": [5, 6, 7]},
        ],
        "regions": regions,
        "summary": {
            "regionCount": len(regions),
            "nationBreakdown": {
                "남한": len(SOUTH_STATIONS),
                "북한": len(NORTH_STATIONS),
            },
            "period": "1991-2020",
            "sourceLabel": "기상청 기후평년값 / 지점정보",
        },
    }
    return dataset


def round_number(value) -> float:
    if value is None:
        return 0.0
    return round(float(value), 1)


def round_coordinate(value) -> float:
    return round(float(value), 4)


def average(values: list[float]) -> float:
    return round(sum(values) / len(values), 1)


def main() -> int:
    dataset = build_dataset()
    OUTPUT_JSON.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.KOREA_CLIMATE_DATA = " + json.dumps(dataset, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"KMA Korea dataset built: {len(dataset['regions'])} regions")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
